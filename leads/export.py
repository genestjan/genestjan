#!/usr/bin/env python3
"""Merge every source into call-ready deliverables, with provenance on each row."""
import json, csv, re
from pathlib import Path
from collections import Counter
import audit

HERE = Path(__file__).parent
OUT = HERE / "output"; (OUT / "by-state").mkdir(parents=True, exist_ok=True)


def load(name, default):
    p = HERE / name
    try:
        return json.loads(p.read_text()) if p.exists() else default
    except Exception:
        return default


base = load("leads_base.json", [])
enr = {e["npi"]: e for e in load("enriched.json", []) if e.get("npi")}
eps = load("endpoint_emails.json", {})
mextra = load("maps_extra.json", {})            # npi -> maps fields
mnew = load("maps_new.json", [])                # maps-only businesses
semail = load("site_emails.json", {})           # key -> email found on site

STATE_NAME = {"CT": "Connecticut", "NY": "New York", "RI": "Rhode Island",
              "MA": "Massachusetts", "MD": "Maryland", "ME": "Maine",
              "VT": "Vermont", "VA": "Virginia"}
BEACHHEAD = {"CT", "NY"}
TYPO = re.compile(r"@(gamil|gmial|gmai|hotmial|yaho|outlok|yahooo|gmail\.co$)", re.I)
NPI_URL = "https://npiregistry.cms.hhs.gov/provider-view/{}"

rows = []
for b in base:
    npi = b.get("npi", "")
    e = enr.get(npi, {})
    mx = mextra.get(npi, {})
    key = f"npi:{npi}"

    website = e.get("website") or mx.get("website") or b.get("website") or ""
    email, esrc, epage = "", "", ""
    if eps.get(npi):
        email, esrc = eps[npi], "NPPES federal registry filing"
        epage = NPI_URL.format(npi)
    elif semail.get(key, {}).get("email"):
        s = semail[key]
        email, esrc, epage = s["email"], "Practice website", s.get("email_page", "")
    elif e.get("email"):
        email = e["email"]
        esrc = ("Practice website (phone-verified)"
                if e.get("verified_by") == "phone" else
                "Practice website (address-verified)")
        epage = e.get("website", "")

    gen = b["tier"] == "PRIMARY"
    prio = "A" if (gen and b["state"] in BEACHHEAD) else ("B" if gen else "C")
    yr = (b.get("last_updated") or "")[:4]
    try:
        yi = int(yr)
    except ValueError:
        yi = 0
    fresh = "Recent" if yi >= 2022 else "Ageing" if yi >= 2015 else "Stale"

    srcs = 1 + (1 if mx.get("maps_url") else 0) + (1 if website else 0)
    rows.append({
        "Priority": prio, "Practice Name": b["practice_name"],
        "City": b["city"], "State": b["state"], "Phone": b["phone"],
        "Owner / Decision Maker": b["owner_name"], "Owner Title": b["owner_title"],
        "Email": email, "Email Source": esrc, "Website": website,
        "Address": b["address"], "ZIP": b["zip"],
        "Practice Type": b["kind"],
        "Record Type": b.get("record_type", "Practice (organisation)"),
        "Target Tier": "Primary (general)" if gen else "Secondary (specialty)",
        "Google Rating": mx.get("rating", ""),
        "Locations On Phone": b.get("sites", 1),
        "Record Updated": yr, "Record Freshness": fresh,
        "Source - Registry": NPI_URL.format(npi) if npi else "",
        "Source - Website": website,
        "Source - Google Maps": mx.get("maps_url", ""),
        "Source - Email Found On": epage,
        "_corro": srcs, "NPI": npi, "Data Flag": "",
    })

# Google-Maps-only businesses (no NPPES match)
for m in mnew:
    key = f"maps:{m.get('maps_url','')}"
    s = semail.get(key, {})
    website = m.get("website", "")
    rows.append({
        "Priority": "D", "Practice Name": m["name"], "City": m["city"],
        "State": m["state"], "Phone": m["phone"], "Owner / Decision Maker": "",
        "Owner Title": "", "Email": s.get("email", ""),
        "Email Source": "Practice website" if s.get("email") else "",
        "Website": website, "Address": m["address"], "ZIP": "",
        "Practice Type": m.get("category") or "Dental",
        "Record Type": "Google Maps listing",
        "Target Tier": "Maps-sourced", "Google Rating": m.get("rating", ""),
        "Locations On Phone": 1, "Record Updated": "", "Record Freshness": "",
        "Source - Registry": "", "Source - Website": website,
        "Source - Google Maps": m.get("maps_url", ""),
        "Source - Email Found On": s.get("email_page", ""),
        "_corro": 1 + (1 if website else 0), "NPI": "", "Data Flag": "",
    })

# ---------- audit every email + phone ----------
print("auditing contact data...")
checks = audit.run(rows, email_key="Email", phone_key="Phone",
                   state_key="State", corro_key="_corro")
for r, c in zip(rows, checks):
    r["Phone Check"] = c["phone_status"]
    r["Phone Note"] = c["phone_detail"]
    r["Email Check"] = c["email_status"]
    r["Email Note"] = c["email_detail"]
    r["Email Type"] = c["email_type"]
    r["Mail Host"] = c["mx_provider"]
    if r["Email"] and TYPO.search(r["Email"]):
        r["Data Flag"] = "CHECK - likely typo in the practice's own filing"
    r.pop("_corro", None)

COLS = ["Priority", "Practice Name", "City", "State", "Phone", "Phone Check",
        "Phone Note", "Owner / Decision Maker", "Owner Title", "Email",
        "Email Check", "Email Type", "Mail Host", "Email Note", "Email Source",
        "Website", "Address", "ZIP", "Practice Type", "Record Type",
        "Target Tier", "Google Rating", "Locations On Phone", "Record Updated",
        "Record Freshness", "Source - Registry", "Source - Website",
        "Source - Google Maps", "Source - Email Found On", "Data Flag", "NPI",
        "Called", "Outcome", "Follow Up"]
for r in rows:
    for c in COLS:
        r.setdefault(c, "")

rows.sort(key=lambda r: (r["Priority"],
                         {"Recent": 0, "Ageing": 1, "Stale": 2, "": 3}[r["Record Freshness"]],
                         r["State"], r["City"], r["Practice Name"]))


def write_csv(path, data):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLS, extrasaction="ignore")
        w.writeheader(); w.writerows(data)


write_csv(OUT / "Hudson_Dental_Leads_MASTER.csv", rows)
write_csv(OUT / "Hudson_Dental_Leads_WITH_EMAIL.csv",
          [r for r in rows if r["Email"] and r["Email Check"] == "DELIVERABLE"])
for st in STATE_NAME:
    sub = [r for r in rows if r["State"] == st]
    if sub:
        write_csv(OUT / "by-state" / f"{st}_{STATE_NAME[st].replace(' ','_')}.csv", sub)

# ---------- workbook ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook(); ws = wb.active; ws.title = "Summary"
HEAD = PatternFill("solid", fgColor="1F3864")
HF = Font(color="FFFFFF", bold=True, size=11)
GOOD = PatternFill("solid", fgColor="C6E0B4")
WARN = PatternFill("solid", fgColor="FFE699")
BAD = PatternFill("solid", fgColor="F8CBAD")


def sheet(ws, data):
    ws.append(COLS)
    for c in ws[1]:
        c.fill = HEAD; c.font = HF
    for r in data:
        ws.append([r.get(c, "") for c in COLS])
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(data)+1}"
    for i, c in enumerate(COLS, 1):
        w = max([len(c) + 2] + [len(str(r.get(c, ""))) + 2 for r in data[:300]])
        ws.column_dimensions[get_column_letter(i)].width = min(w, 40)
    ec, pc = COLS.index("Email Check") + 1, COLS.index("Phone Check") + 1
    for row in ws.iter_rows(min_row=2, max_row=len(data) + 1):
        if row[0].value == "A":
            row[0].fill = GOOD
        row[0].alignment = Alignment(horizontal="center")
        for idx in (ec, pc):
            v = row[idx - 1].value
            if v in ("DELIVERABLE", "VALID"):
                row[idx - 1].fill = GOOD
            elif v in ("RISKY", "CHECK"):
                row[idx - 1].fill = WARN
            elif v in ("DEAD", "INVALID", "REJECT"):
                row[idx - 1].fill = BAD


emails = [r for r in rows if r["Email"]]
deliv = [r for r in emails if r["Email Check"] == "DELIVERABLE"]
sites = [r for r in rows if r["Website"]]
maps_rows = [r for r in rows if r["Source - Google Maps"]]
summary = [
    ["HUDSON HECTOR - DENTAL PRACTICE CALL LIST", ""],
    ["Built", "30 August 2026"],
    ["Sources", "CMS NPPES federal provider registry + Google Maps + practice websites"],
    ["", ""],
    ["TOTALS", ""],
    ["Total leads", len(rows)],
    ["  With a phone number", sum(1 for r in rows if r["Phone"])],
    ["  With a named owner / decision maker", sum(1 for r in rows if r["Owner / Decision Maker"])],
    ["  With a website", len(sites)],
    ["  With an email", len(emails)],
    ["  Corroborated by a Google Maps listing", len(maps_rows)],
    ["", ""],
    ["CONTACT AUDIT - PHONES", ""],
]
for k, n in Counter(r["Phone Check"] for r in rows).most_common():
    summary.append([f"  {k or '(no phone)'}", n])
summary += [["", ""], ["CONTACT AUDIT - EMAILS", ""]]
for k, n in Counter(r["Email Check"] for r in rows if r["Email"]).most_common():
    summary.append([f"  {k}", n])
summary += [
    ["", ""],
    ["CALL PRIORITY", ""],
    ["  A - general dentistry, CT + NY", sum(1 for r in rows if r["Priority"] == "A")],
    ["  B - general dentistry, other states", sum(1 for r in rows if r["Priority"] == "B")],
    ["  C - specialty practices", sum(1 for r in rows if r["Priority"] == "C")],
    ["  D - Google Maps only (no registry match)", sum(1 for r in rows if r["Priority"] == "D")],
    ["", ""], ["BY STATE", ""],
]
for st, n in Counter(r["State"] for r in rows).most_common():
    summary.append([f"  {STATE_NAME.get(st, st)} ({st})", n])
for r in summary:
    ws.append(r)
ws["A1"].font = Font(bold=True, size=14, color="1F3864")
ws.column_dimensions["A"].width = 48; ws.column_dimensions["B"].width = 62
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):
    v = row[0].value
    if v and str(v).isupper():
        row[0].font = Font(bold=True)

sheet(wb.create_sheet("All Leads"), rows)
if deliv:
    sheet(wb.create_sheet("Email - Deliverable"), deliv)
for st in STATE_NAME:
    sub = [r for r in rows if r["State"] == st]
    if sub:
        sheet(wb.create_sheet(STATE_NAME[st][:28]), sub)
wb.save(OUT / "Hudson_Dental_Leads.xlsx")

print(f"\nleads           {len(rows)}")
print(f"emails          {len(emails)} (deliverable {len(deliv)})")
print(f"websites        {len(sites)}")
print(f"maps-sourced    {len(maps_rows)}")
print(f"-> {OUT}")
